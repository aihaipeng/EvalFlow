from pathlib import Path
from shutil import copy2

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


TARGET = Path(r"C:\Users\Administrator\Desktop\testcases.xlsx")
BACKUP = TARGET.with_name("testcases.before-datacenter-demo.xlsx")
SCENARIOS = [
    ("温控系统失效，机柜温度持续升高", "COOLING_FAILURE", "P1", "SWITCH_TO_BACKUP_COOLING"),
    ("双路供电中的主路电压异常", "POWER_SUPPLY_FAILURE", "P1", "TRANSFER_TO_BACKUP_POWER"),
    ("核心交换链路出现持续丢包", "NETWORK_PACKET_LOSS", "P2", "REROUTE_NETWORK_TRAFFIC"),
    ("存储阵列 I/O 延迟急剧上升", "DISK_IO_SATURATION", "P2", "MIGRATE_IO_WORKLOAD"),
    ("应用进程内存持续增长并触发 OOM", "MEMORY_LEAK", "P2", "RESTART_AND_PATCH_SERVICE"),
]


def main() -> None:
    if TARGET.exists() and not BACKUP.exists():
        copy2(TARGET, BACKUP)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "机房故障诊断"
    headers = [
        "case_id", "query", "alarm_type", "datacenter", "rack", "host",
        "expected_root_cause", "expected_risk_level", "expected_action", "rule_description",
    ]
    sheet.append(headers)
    for index in range(1, 101):
        description, root_cause, risk, action = SCENARIOS[(index - 1) % len(SCENARIOS)]
        sheet.append([
            f"CASE-{index:03d}",
            f"请分析 DC-SH-01 机房第 {index:03d} 号告警：{description}，给出根因、风险等级和处置动作。",
            root_cause,
            "DC-SH-01",
            f"RACK-{(index % 20) + 1:02d}",
            f"srv-{(index % 50) + 1:03d}",
            root_cause,
            risk,
            action,
            f"根因={root_cause}; 风险={risk}; 动作={action}",
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    widths = [14, 70, 26, 16, 14, 14, 28, 22, 34, 64]
    for column, width in zip(sheet.columns, widths):
        sheet.column_dimensions[column[0].column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(TARGET)

    verified = load_workbook(TARGET, read_only=True, data_only=True)["机房故障诊断"]
    assert verified.max_row == 101
    assert verified.max_column == len(headers)
    print(f"created={TARGET} cases={verified.max_row - 1} columns={verified.max_column}")


if __name__ == "__main__":
    main()
