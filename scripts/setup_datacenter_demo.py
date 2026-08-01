"""Create the enterprise datacenter diagnostic test set, workflow, and batch run."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import Request, urlopen
from uuid import uuid4

from openpyxl import load_workbook


API_BASE = "http://127.0.0.1:8010/api"
EXCEL_PATH = Path.home() / "Desktop" / "testcases.xlsx"
PROVIDER_ID = "0c4fc72ef8f44eea99d4c0a532ef9e3c"
MODEL_NAME = "deepseek-v4-pro"
ARTIFACT_PATH = Path("run_storage/datacenter-demo-setup.json")


def request(method: str, path: str, body: dict | None = None) -> dict:
    payload = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = Request(
        API_BASE + path,
        data=payload,
        method=method,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urlopen(req, timeout=60) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{method} {path} failed: HTTP {exc.code}: {detail}") from exc


def load_cases() -> tuple[list[str], list[dict]]:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook["机房故障诊断"]
    rows = sheet.iter_rows(values_only=True)
    columns = [str(value) for value in next(rows)]
    cases = []
    for row in rows:
        values = {columns[index]: "" if value is None else str(value) for index, value in enumerate(row)}
        cases.append({"id": values["case_id"], "values": values})
    workbook.close()
    if len(cases) != 100:
        raise RuntimeError(f"expected 100 cases, got {len(cases)}")
    return columns, cases


def node_id() -> str:
    return str(uuid4())


def edge(source: str, target: str) -> dict:
    return {"id": str(uuid4()), "source_node_id": source, "target_node_id": target}


def retry(timeout: float = 180) -> dict:
    return {
        "timeout_seconds": timeout,
        "max_attempts": 1,
        "retry_interval_seconds": 2,
        "delay_seconds": 0,
    }


def placed(node: dict, x: float, y: float) -> dict:
    return {"node": node, "position_x": x, "position_y": y}


def llm_node(node: str, name: str, output: str, system: str, user: str) -> dict:
    return {
        "id": node,
        "type": "LLM",
        "name": name,
        "description": "使用 DeepSeek 对独立机房诊断校验点进行复核",
        "model": {"provider_id": PROVIDER_ID, "model_name": MODEL_NAME},
        "context": {
            "messages": [
                {"role": "SYSTEM", "content": system},
                {"role": "USER", "content": user},
            ]
        },
        "generation": {
            "parameters": {"temperature": 0, "max_output_tokens": 1200},
            "parameters_text": "",
        },
        "execution": retry(240),
        "outputs": [
            {
                "name": output,
                "type": "string",
                "source": "result.output",
            }
        ],
    }


def build_workflow(name: str) -> dict:
    ids = {key: node_id() for key in (
        "start", "http", "parse", "root_llm", "risk_llm", "action_llm",
        "root_script", "risk_script", "action_script", "aggregate", "end",
    )}
    start_inputs = [
        {"name": key, "type": "string", "value": ""}
        for key in (
            "case_id", "query", "alarm_type", "datacenter", "rack", "host",
            "expected_root_cause", "expected_risk_level", "expected_action",
            "rule_description",
        )
    ]
    parse_script = """payload = context[\"diagnostic_payload\"]
diagnosis = payload[\"diagnosis\"]
metrics = payload[\"metrics\"]
logs = payload[\"logs\"]
parsed_checks = {
    \"case_id\": payload[\"case_id\"],
    \"location\": {\"datacenter\": payload[\"datacenter\"], \"rack\": payload[\"rack\"], \"host\": payload[\"host\"]},
    \"root_cause\": diagnosis[\"root_cause\"],
    \"risk_level\": diagnosis[\"risk_level\"],
    \"recommended_action\": diagnosis[\"recommended_action\"],
    \"confidence\": diagnosis[\"confidence\"],
    \"affected_hosts\": payload[\"impact\"][\"affected_hosts\"],
    \"affected_services\": payload[\"impact\"][\"affected_services\"],
    \"metric_count\": len(metrics),
    \"error_log_count\": sum(1 for item in logs if item[\"level\"] == \"ERROR\"),
    \"max_temperature\": max(item[\"temperature\"] for item in metrics),
    \"max_packet_loss\": max(item[\"packet_loss\"] for item in metrics),
    \"max_disk_latency_ms\": max(item[\"disk_latency_ms\"] for item in metrics),
    \"recent_error_samples\": [item[\"message\"] for item in logs if item[\"level\"] == \"ERROR\"][:20],
}
"""
    aggregate_script = """checks = context[\"parsed_checks\"]
root_cause_match = checks[\"root_cause\"] == context[\"expected_root_cause\"]
risk_level_match = checks[\"risk_level\"] == context[\"expected_risk_level\"]
action_match = checks[\"recommended_action\"] == context[\"expected_action\"]
diagnostic_summary = {
    \"case_id\": context[\"case_id\"],
    \"observed\": {
        \"root_cause\": checks[\"root_cause\"],
        \"risk_level\": checks[\"risk_level\"],
        \"recommended_action\": checks[\"recommended_action\"],
    },
    \"expected\": {
        \"root_cause\": context[\"expected_root_cause\"],
        \"risk_level\": context[\"expected_risk_level\"],
        \"recommended_action\": context[\"expected_action\"],
    },
    \"reviews\": {
        \"root_cause\": context[\"root_review_record\"],
        \"risk\": context[\"risk_review_record\"],
        \"action\": context[\"action_review_record\"],
    },
    \"all_checks_passed\": root_cause_match and risk_level_match and action_match,
}
"""
    nodes = [
        placed({"id": ids["start"], "type": "START", "name": "输入机房告警", "description": "接收测试集字段", "inputs": start_inputs}, 40, 280),
        placed({
            "id": ids["http"], "type": "HTTP", "name": "调用企业诊断智能体", "description": "获取约 210KB 的诊断响应",
            "request": {
                "method": "POST", "url": "http://127.0.0.1:9000/datacenter/diagnose", "follow_redirects": True,
                "headers": [{"key": "Content-Type", "value": "application/json"}], "params": [],
                "body": {"type": "raw", "content": None, "template_text": '{"case_id": ${case_id}, "query": ${query}}'},
            },
            "network": {"proxy": {"mode": "DIRECT", "url": None, "username": None, "password": None}, "verify_ssl": True},
            "response": {"mode": "JSON", "success_statuses": ["200-299"]},
            "execution": {**retry(60), "retry_non_idempotent": True, "retry_statuses": [408, 429, 500, 502, 503, 504]},
            "outputs": [{"name": "diagnostic_payload", "type": "object", "source": "response.body.data"}],
        }, 300, 280),
        placed({
            "id": ids["parse"], "type": "SCRIPT", "name": "解析大响应并拆分校验点", "description": "压缩指标、日志和拓扑信息",
            "script": parse_script, "execution": retry(30),
            "outputs": [{"name": "parsed_checks", "type": "object", "source": "parsed_checks"}],
        }, 610, 280),
        placed(llm_node(ids["root_llm"], "LLM 根因复核", "root_review", "你是企业数据中心一级故障诊断专家。基于证据复核根因，明确说明证据链与置信度，不要省略推理结论。", "诊断摘要：${parsed_checks}\n告警：${query}\n预期根因：${expected_root_cause}\n请判断根因是否一致并给出处置风险。"), 930, 40),
        placed(llm_node(ids["risk_llm"], "LLM 风险复核", "risk_review", "你是企业数据中心 SRE 风险评估专家。评估业务影响、升级等级和潜在扩散范围。", "诊断摘要：${parsed_checks}\n规则：${rule_description}\n预期风险等级：${expected_risk_level}\n请给出完整风险复核。"), 930, 280),
        placed(llm_node(ids["action_llm"], "LLM 处置方案复核", "action_review", "你是企业机房应急处置负责人。评估操作是否与故障匹配、是否可回滚以及执行顺序。", "诊断摘要：${parsed_checks}\n预期动作：${expected_action}\n请复核动作并列出执行前检查和回滚条件。"), 930, 520),
        placed({
            "id": ids["root_script"], "type": "SCRIPT", "name": "标准化根因复核", "description": "保留模型审计文本",
            "script": 'text = context["root_review"]\nroot_review_record = {"received": bool(text.strip()), "text": text}', "execution": retry(20),
            "outputs": [{"name": "root_review_record", "type": "object", "source": "root_review_record"}],
        }, 1260, 40),
        placed({
            "id": ids["risk_script"], "type": "SCRIPT", "name": "标准化风险复核", "description": "保留模型审计文本",
            "script": 'text = context["risk_review"]\nrisk_review_record = {"received": bool(text.strip()), "text": text}', "execution": retry(20),
            "outputs": [{"name": "risk_review_record", "type": "object", "source": "risk_review_record"}],
        }, 1260, 280),
        placed({
            "id": ids["action_script"], "type": "SCRIPT", "name": "标准化处置复核", "description": "保留模型审计文本",
            "script": 'text = context["action_review"]\naction_review_record = {"received": bool(text.strip()), "text": text}', "execution": retry(20),
            "outputs": [{"name": "action_review_record", "type": "object", "source": "action_review_record"}],
        }, 1260, 520),
        placed({
            "id": ids["aggregate"], "type": "SCRIPT", "name": "汇总诊断与规则结果", "description": "汇总三个 LLM 分支并生成调度校验字段",
            "script": aggregate_script, "execution": retry(30),
            "outputs": [
                {"name": "root_cause_match", "type": "boolean", "source": "root_cause_match"},
                {"name": "risk_level_match", "type": "boolean", "source": "risk_level_match"},
                {"name": "action_match", "type": "boolean", "source": "action_match"},
                {"name": "diagnostic_summary", "type": "object", "source": "diagnostic_summary"},
            ],
        }, 1580, 280),
        placed({
            "id": ids["end"], "type": "END", "name": "输出诊断结论", "description": "输出最终校验与可审计详情",
            "outputs": [
                {"name": "root_cause_match", "type": "boolean", "source": "root_cause_match"},
                {"name": "risk_level_match", "type": "boolean", "source": "risk_level_match"},
                {"name": "action_match", "type": "boolean", "source": "action_match"},
                {"name": "diagnostic_summary", "type": "object", "source": "diagnostic_summary"},
            ],
        }, 1880, 280),
    ]
    edges = [
        edge(ids["start"], ids["http"]), edge(ids["http"], ids["parse"]),
        edge(ids["parse"], ids["root_llm"]), edge(ids["parse"], ids["risk_llm"]), edge(ids["parse"], ids["action_llm"]),
        edge(ids["root_llm"], ids["root_script"]), edge(ids["risk_llm"], ids["risk_script"]), edge(ids["action_llm"], ids["action_script"]),
        edge(ids["root_script"], ids["aggregate"]), edge(ids["risk_script"], ids["aggregate"]), edge(ids["action_script"], ids["aggregate"]),
        edge(ids["aggregate"], ids["end"]),
    ]
    return {"name": name, "description": "企业机房故障诊断：大响应解析、三路 DeepSeek 专项复核与规则汇总", "nodes": nodes, "edges": edges}


def main() -> None:
    columns, cases = load_cases()
    suffix = datetime.now().strftime("%Y%m%d-%H%M%S")
    test_set_name = f"企业机房故障诊断-100例-{suffix}"
    workflow_name = f"企业机房智能故障诊断-{suffix}"
    batch_name = f"机房故障诊断批量验证-{suffix}"

    test_set = request("POST", "/test-sets", {
        "name": test_set_name,
        "description": "100 条企业机房告警场景，包含根因、风险等级、处置动作和可读规则",
        "columns": columns,
        "cases": cases,
    })["test_set"]
    workflow = request("POST", "/workflows", build_workflow(workflow_name))["workflow"]
    workflow_id = workflow["workflow"]["id"]
    variables = [
        {"source": "TEST_SET", "key": column, "value": column, "type": "string"}
        for column in columns
    ]
    batch = request("POST", "/batch-runs", {
        "name": batch_name,
        "description": "调用本机企业诊断智能体，使用三路 deepseek-v4-pro 复核 100 条机房告警",
        "test_set_id": test_set["id"],
        "workflow_id": workflow_id,
        "variables": variables,
        "case_concurrency": 2,
        "call_order": "SEQUENTIAL",
        "evaluation_rules": [
            {"name": "根因", "result_path": "context.root_cause_match", "operator": "EQ", "expected_value": "true", "type": "boolean"},
            {"name": "风险等级", "result_path": "context.risk_level_match", "operator": "EQ", "expected_value": "true", "type": "boolean"},
            {"name": "处置动作", "result_path": "context.action_match", "operator": "EQ", "expected_value": "true", "type": "boolean"},
        ],
        "case_display_column": "query",
        "rule_display_column": "rule_description",
    })["batch"]
    artifact = {
        "created_at": datetime.now().astimezone().isoformat(),
        "test_set": {"id": test_set["id"], "name": test_set_name, "cases": len(cases)},
        "workflow": {"id": workflow_id, "name": workflow_name, "nodes": len(workflow["node_models"])},
        "batch": {"id": batch["id"], "name": batch_name, "status": batch["status"]},
    }
    ARTIFACT_PATH.parent.mkdir(parents=True, exist_ok=True)
    ARTIFACT_PATH.write_text(json.dumps(artifact, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(artifact, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
