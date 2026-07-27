import math
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


CASE_ID_HEADERS = {"case_id", "case id", "用例id", "用例编号"}
QUESTION_HEADERS = {"question", "问题"}


def _as_text(value) -> str:
    """把 Excel 单元格值转换为去除首尾空白的文本。"""
    return "" if value is None else str(value).strip()


@dataclass(frozen=True)
class TestCase:
    """从 Excel 读取的一条 Web 管理用测试用例。"""

    case_id: str
    question: str
    row_number: int


def read_test_cases_from_sheet(sheet) -> list[TestCase]:
    """从一个已打开的工作表读取有效用例并保留 Excel 行号。"""
    rows = sheet.iter_rows(min_col=1, max_col=2, values_only=True)
    testcases: list[TestCase] = []
    seen_ids: set[str] = set()
    for row_number, row in enumerate(rows, start=1):
        values = list(row) + [None] * (2 - len(row))
        case_id = _as_text(values[0])
        question = _as_text(values[1])

        if row_number == 1 and (
            case_id.casefold() in CASE_ID_HEADERS
            or question.casefold() in QUESTION_HEADERS
        ):
            continue
        if not case_id or not question:
            continue
        if case_id in seen_ids:
            continue
        seen_ids.add(case_id)
        testcases.append(
            TestCase(
                case_id=case_id,
                question=question,
                row_number=row_number,
            )
        )
    return testcases


class ExcelCaseRepository:
    """读取测试用例 Excel。

    当前 Web 项目只支持固定两列输入格式：``case_id | question``。
    第三列及之后允许存在历史结果或人工备注，但不会参与用例读取。
    """

    def __init__(self, file_path: str | Path, sheet_name: str = "Sheet1"):
        """绑定一个测试用例工作簿及目标工作表。"""
        self.path = Path(file_path)
        self.sheet_name = sheet_name

    def _select_sheet(self, workbook):
        """选择指定工作表。"""
        if self.sheet_name in workbook.sheetnames:
            return workbook[self.sheet_name]
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet 不存在: {self.sheet_name}。可用: {available}")

    def read_cases(self) -> list[TestCase]:
        """读取工作簿中的有效测试用例。

        空行、表头、空 ID、空问题和重复 ID 会被忽略。
        """
        if not self.path.is_file():
            raise FileNotFoundError(f"Excel 文件不存在: {self.path}")

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = self._select_sheet(workbook)
            return read_test_cases_from_sheet(sheet)
        finally:
            workbook.close()


@dataclass(frozen=True)
class ExcelSheetRow:
    """批量运行使用的一条严格表头行记录。"""

    row_number: int
    values: dict[str, Any]


@dataclass(frozen=True)
class ExcelSheetSnapshot:
    """批量运行创建前读取的 Sheet 表头与非空数据行。"""

    headers: tuple[str, ...]
    rows: tuple[ExcelSheetRow, ...]
    header_mode: str


def _json_cell_value(value: Any, *, coordinate: str) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError(f"单元格 {coordinate} 包含非有限数值")
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    raise ValueError(f"单元格 {coordinate} 的类型不支持批量执行: {type(value).__name__}")


class ExcelSheetRepository:
    """按首行表头读取批量运行输入，不应用用例过滤或字段映射。"""

    def __init__(
        self,
        file_path: str | Path,
        sheet_name: str,
        header_mode: str = "AUTO",
    ):
        self.path = Path(file_path)
        self.sheet_name = sheet_name
        self.header_mode = header_mode.upper()
        if self.header_mode not in {"AUTO", "HEADER", "DATA"}:
            raise ValueError("首行模式必须是 AUTO、HEADER 或 DATA")

    def read_snapshot(self) -> ExcelSheetSnapshot:
        if not self.path.is_file():
            raise FileNotFoundError(f"Excel 文件不存在: {self.path}")

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            if self.sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise ValueError(f"Sheet 不存在: {self.sheet_name}。可用: {available}")
            sheet = workbook[self.sheet_name]
            raw_rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            if not raw_rows:
                raise ValueError("Sheet 为空，没有可执行数据")

            first = [_as_text(value) for value in raw_rows[0]]
            auto_header = bool(first) and (
                first[0].casefold() in CASE_ID_HEADERS
                or (len(first) > 1 and first[1].casefold() in QUESTION_HEADERS)
            )
            resolved_mode = (
                "HEADER"
                if self.header_mode == "HEADER" or (self.header_mode == "AUTO" and auto_header)
                else "DATA"
            )
            if resolved_mode == "HEADER":
                last_header = max(
                    (index for index, header in enumerate(first) if header),
                    default=-1,
                )
                if last_header < 0:
                    raise ValueError("Sheet 第一行必须提供至少一个非空表头")
                headers = tuple(first[: last_header + 1])
                if any(not header for header in headers):
                    raise ValueError("Sheet 有效表头区间中不能包含空列名")
                for raw_row in raw_rows[1:]:
                    if any(
                        value is not None and (not isinstance(value, str) or value.strip())
                        for value in raw_row[len(headers) :]
                    ):
                        raise ValueError("Sheet 存在没有表头的数据列")
                data_rows = raw_rows[1:]
                first_row_number = 2
            else:
                width = max(
                    (
                        index + 1
                        for row in raw_rows
                        for index, value in enumerate(row)
                        if value is not None and (not isinstance(value, str) or value.strip())
                    ),
                    default=0,
                )
                if width == 0:
                    raise ValueError("Sheet 没有可执行数据")
                headers = tuple(
                    "case_id" if index == 1 else "question" if index == 2 else f"column_{index}"
                    for index in range(1, width + 1)
                )
                data_rows = raw_rows
                first_row_number = 1
            duplicates = sorted({header for header in headers if headers.count(header) > 1})
            if duplicates:
                raise ValueError(f"Sheet 表头不能重复: {', '.join(duplicates)}")

            rows: list[ExcelSheetRow] = []
            for row_number, raw_row in enumerate(data_rows, start=first_row_number):
                values = list(raw_row) + [None] * (len(headers) - len(raw_row))
                values = values[: len(headers)]
                if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
                    continue
                normalized = {
                    header: _json_cell_value(
                        values[index],
                        coordinate=f"{get_column_letter(index + 1)}{row_number}",
                    )
                    for index, header in enumerate(headers)
                }
                rows.append(ExcelSheetRow(row_number=row_number, values=normalized))
            return ExcelSheetSnapshot(
                headers=headers,
                rows=tuple(rows),
                header_mode=resolved_mode,
            )
        finally:
            workbook.close()
